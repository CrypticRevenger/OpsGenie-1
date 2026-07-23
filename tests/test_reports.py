"""Period-scoped report builders — party ledger, payment/sales/purchase
registers, day book, outstanding aging — plus the shared period resolver,
party lookup, and the REPORTS registry tying them together.

    uv run alembic upgrade head
    uv run pytest tests/test_reports.py -v
"""

from __future__ import annotations

import io
import uuid
from datetime import date, timedelta
from decimal import Decimal

import pytest
from app.models.company import Company
from app.models.dealer import Dealer
from app.models.invoice import Invoice, InvoiceDirection, InvoiceSource, InvoiceStatus
from app.models.invoice_item import InvoiceItem
from app.models.payment import Payment, PaymentSource
from app.models.supplier import Supplier
from app.services.party_lookup import find_party, get_party_by_id
from app.services.party_outstanding import calculate_party_outstanding
from app.services.reports import aging, ledger, registers
from app.services.reports.period import InvalidPeriodError, resolve_period
from app.services.reports.registry import REPORTS, ReportContext
from app.services.snapshot import business_now
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


def _unique_phone() -> str:
    return f"+91{uuid.uuid4().int % 10_000_000_000:010d}"


async def _make_company(db: AsyncSession, name: str = "Report Test Co") -> Company:
    company = Company(business_name=name, owner_name="Owner", whatsapp_number=_unique_phone())
    db.add(company)
    await db.commit()
    await db.refresh(company)
    return company


def _invoice(
    company: Company,
    *,
    direction: InvoiceDirection,
    dealer: Dealer | None = None,
    supplier: Supplier | None = None,
    invoice_date: date,
    due_date: date,
    total: Decimal,
    status: InvoiceStatus = InvoiceStatus.Pending,
) -> Invoice:
    return Invoice(
        company_id=company.id,
        invoice_number=f"T-{uuid.uuid4().hex[:10]}",
        direction=direction,
        dealer_id=dealer.id if dealer else None,
        supplier_id=supplier.id if supplier else None,
        invoice_date=invoice_date,
        due_date=due_date,
        subtotal=total,
        gst_amount=Decimal("0"),
        total_amount=total,
        status=status,
        source=InvoiceSource.whatsapp,
    )


def _payment(company: Company, invoice: Invoice, *, amount: Decimal, payment_date: date) -> Payment:
    return Payment(
        company_id=company.id,
        invoice_id=invoice.id,
        amount=amount,
        payment_date=payment_date,
        source=PaymentSource.whatsapp,
    )


def _load(wb_bytes: bytes):
    return load_workbook(io.BytesIO(wb_bytes))


def _find_row(rows: list[tuple], *, col: int, value: object) -> tuple:
    return next(r for r in rows if r[col] == value)


# ── resolve_period ───────────────────────────────────────────────────────────


def test_resolve_period_month_shortcut() -> None:
    period = resolve_period(month_str="2026-07")
    assert period.from_date == date(2026, 7, 1)
    assert period.to_date == date(2026, 7, 31)
    assert period.label == "July 2026"


def test_resolve_period_explicit_range() -> None:
    period = resolve_period(from_str="2026-01-01", to_str="2026-01-15")
    assert period.from_date == date(2026, 1, 1)
    assert period.to_date == date(2026, 1, 15)


def test_resolve_period_open_ended_from() -> None:
    period = resolve_period(from_str="2026-01-01")
    assert period.from_date == date(2026, 1, 1)
    assert period.to_date is None


def test_resolve_period_defaults_to_current_month_when_requested() -> None:
    period = resolve_period(default_to_current_month=True, today=date(2026, 7, 19))
    assert period.from_date == date(2026, 7, 1)
    assert period.to_date == date(2026, 7, 31)


def test_resolve_period_all_time_by_default() -> None:
    period = resolve_period()
    assert period.from_date is None
    assert period.to_date is None
    assert period.label == "All time"


def test_resolve_period_invalid_month_raises() -> None:
    with pytest.raises(InvalidPeriodError):
        resolve_period(month_str="not-a-month")


def test_resolve_period_from_after_to_raises() -> None:
    with pytest.raises(InvalidPeriodError):
        resolve_period(from_str="2026-02-01", to_str="2026-01-01")


# ── party_lookup ─────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_find_party_matches_dealer_over_supplier(db: AsyncSession) -> None:
    company = await _make_company(db)
    db.add(Dealer(company_id=company.id, name="Ram Traders"))
    db.add(Supplier(company_id=company.id, name="Ram Traders Supply Co"))
    await db.commit()

    match = await find_party(db, company.id, "ram")
    assert match is not None
    party, direction = match
    assert direction == "receivable"
    assert party.name == "Ram Traders"


@pytest.mark.asyncio
async def test_find_party_matches_supplier_when_no_dealer(db: AsyncSession) -> None:
    company = await _make_company(db)
    db.add(Supplier(company_id=company.id, name="Acme Supply"))
    await db.commit()

    match = await find_party(db, company.id, "acme")
    assert match is not None
    party, direction = match
    assert direction == "payable"
    assert party.name == "Acme Supply"


@pytest.mark.asyncio
async def test_find_party_no_match_returns_none(db: AsyncSession) -> None:
    company = await _make_company(db)
    assert await find_party(db, company.id, "nobody") is None


@pytest.mark.asyncio
async def test_get_party_by_id_scoped_to_company(db: AsyncSession) -> None:
    company_a = await _make_company(db, name="Party Scope A")
    company_b = await _make_company(db, name="Party Scope B")
    dealer = Dealer(company_id=company_a.id, name="Ram Traders")
    db.add(dealer)
    await db.commit()
    await db.refresh(dealer)

    match = await get_party_by_id(db, company_a.id, dealer.id)
    assert match is not None
    assert match[1] == "receivable"
    assert await get_party_by_id(db, company_b.id, dealer.id) is None


# ── Party ledger ─────────────────────────────────────────────────────────────


async def _seed_ledger_dealer(db: AsyncSession, company: Company) -> Dealer:
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    db.add(dealer)
    await db.flush()

    # June: fully paid — nets to zero, must still appear in an all-time ledger.
    june_invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 6, 5),
        due_date=date(2026, 6, 20),
        total=Decimal("1000.00"),
        status=InvoiceStatus.Paid,
    )
    db.add(june_invoice)
    await db.flush()
    db.add(
        _payment(company, june_invoice, amount=Decimal("1000.00"), payment_date=date(2026, 6, 10))
    )

    # July: partially paid — the genuinely open balance.
    july_invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 5),
        due_date=date(2026, 7, 20),
        total=Decimal("2000.00"),
        status=InvoiceStatus.Partially_Paid,
    )
    db.add(july_invoice)
    await db.flush()
    db.add(
        _payment(company, july_invoice, amount=Decimal("500.00"), payment_date=date(2026, 7, 10))
    )

    await db.commit()
    await db.refresh(dealer)
    return dealer


@pytest.mark.asyncio
async def test_ledger_all_time_closing_matches_calculate_party_outstanding(
    db: AsyncSession,
) -> None:
    company = await _make_company(db)
    dealer = await _seed_ledger_dealer(db, company)

    period = resolve_period()  # all-time
    wb = _load(await ledger.build_party_ledger_workbook(db, company, dealer, "receivable", period))
    rows = list(wb["Ledger"].iter_rows(values_only=True))

    opening_row = _find_row(rows, col=1, value="Opening Balance")
    closing_row = _find_row(rows, col=1, value="Closing Balance")
    assert opening_row[5] == 0

    outstanding = await calculate_party_outstanding(db, direction="receivable", party_id=dealer.id)
    assert Decimal(str(closing_row[5])) == outstanding
    assert outstanding == Decimal("1500.00")  # 2000 - 500; June nets to zero


@pytest.mark.asyncio
async def test_ledger_month_scoped_opening_balance_carries_prior_activity(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = await _seed_ledger_dealer(db, company)

    period = resolve_period(month_str="2026-07")
    wb = _load(await ledger.build_party_ledger_workbook(db, company, dealer, "receivable", period))
    rows = list(wb["Ledger"].iter_rows(values_only=True))

    # June's invoice + payment net to zero, so July's opening balance is 0
    # even though June's transactions happened and aren't shown this month.
    opening_row = _find_row(rows, col=1, value="Opening Balance")
    closing_row = _find_row(rows, col=1, value="Closing Balance")
    assert opening_row[5] == 0
    assert closing_row[5] == 1500

    kinds = [r[1] for r in rows if r[1] in ("Invoice", "Payment")]
    assert kinds == ["Invoice", "Payment"]  # only July's rows, June excluded


@pytest.mark.asyncio
async def test_ledger_meta_block_reports_name_period_and_row_count(db: AsyncSession) -> None:
    company = await _make_company(db, name="Meta Block Co")
    dealer = await _seed_ledger_dealer(db, company)

    period = resolve_period(month_str="2026-07")
    wb = _load(await ledger.build_party_ledger_workbook(db, company, dealer, "receivable", period))
    rows = list(wb["Ledger"].iter_rows(values_only=True))

    assert _find_row(rows, col=0, value="Report")[1] == "Ledger — Ram Traders"
    assert _find_row(rows, col=0, value="Company")[1] == "Meta Block Co"
    assert _find_row(rows, col=0, value="Period")[1] == "July 2026"
    assert _find_row(rows, col=0, value="Rows")[1] == 2  # July invoice + July payment


@pytest.mark.asyncio
async def test_ledger_freeze_and_filter_anchor_to_the_real_header_row(db: AsyncSession) -> None:
    # Regression: xlsx_common.write_header hardcoded freeze_panes="A2" and
    # auto_filter row 1, which was only ever correct for a sheet with nothing
    # written above the header. Every report in this package calls
    # write_meta_block first (5 label/value rows + 1 blank = 6 rows), so the
    # real header lands on row 7 — freeze/filter must anchor there, not row 2/1.
    company = await _make_company(db)
    dealer = await _seed_ledger_dealer(db, company)

    period = resolve_period(month_str="2026-07")
    wb = _load(await ledger.build_party_ledger_workbook(db, company, dealer, "receivable", period))
    ws = wb["Ledger"]
    assert ws.freeze_panes == "A8"
    assert ws.auto_filter.ref == "A7:F7"
    header_row = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))
    assert header_row == ("Date", "Type", "Reference", "Debit", "Credit", "Balance")


@pytest.mark.asyncio
async def test_ledger_pdf_starts_with_pdf_magic_bytes(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = await _seed_ledger_dealer(db, company)
    out = await ledger.build_party_ledger_pdf(db, company, dealer, "receivable", resolve_period())
    assert out[:4] == b"%PDF"


# ── Registers ────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_sales_register_rate_wise_summary_matches_line_rows(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Ram Traders", gst_number="21ABCDE1234F1Z5")
    db.add(dealer)
    await db.flush()

    invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 10),
        due_date=date(2026, 7, 25),
        total=Decimal("1162.00"),
        status=InvoiceStatus.Pending,
    )
    db.add(invoice)
    await db.flush()
    db.add(
        InvoiceItem(
            invoice_id=invoice.id,
            description="Rice",
            quantity=Decimal("10"),
            unit_price=Decimal("100"),
            line_total=Decimal("1000"),
            gst_rate=Decimal("5"),
            gst_amount=Decimal("50"),
        )
    )
    db.add(
        InvoiceItem(
            invoice_id=invoice.id,
            description="Oil",
            quantity=Decimal("1"),
            unit_price=Decimal("100"),
            line_total=Decimal("100"),
            gst_rate=Decimal("12"),
            gst_amount=Decimal("12"),
        )
    )
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_sales_register_workbook(db, company, period))

    reg_rows = list(wb["Sales Register"].iter_rows(values_only=True))
    reg_total = _find_row(reg_rows, col=0, value="TOTAL")
    assert reg_total[5] == 1100  # total taxable value
    assert reg_total[7] == 62  # total GST
    assert any(r[3] == "21ABCDE1234F1Z5" for r in reg_rows)  # GSTIN column populated

    summary_rows = list(wb["Rate-wise Summary"].iter_rows(values_only=True))
    summary_data = summary_rows[1:-1]  # header, then one row per rate, then TOTAL
    assert sum(r[1] for r in summary_data) == reg_total[5]
    assert sum(r[2] for r in summary_data) == reg_total[7]


@pytest.mark.asyncio
async def test_purchase_register_handles_supplier_with_no_gstin(db: AsyncSession) -> None:
    company = await _make_company(db)
    supplier = Supplier(company_id=company.id, name="Acme Supply")
    db.add(supplier)
    await db.flush()

    invoice = _invoice(
        company,
        direction=InvoiceDirection.payable,
        supplier=supplier,
        invoice_date=date(2026, 7, 12),
        due_date=date(2026, 7, 27),
        total=Decimal("530"),
        status=InvoiceStatus.Pending,
    )
    db.add(invoice)
    await db.flush()
    db.add(
        InvoiceItem(
            invoice_id=invoice.id,
            description="Packaging",
            quantity=Decimal("1"),
            unit_price=Decimal("500"),
            line_total=Decimal("500"),
            gst_rate=Decimal("6"),
            gst_amount=Decimal("30"),
        )
    )
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_purchase_register_workbook(db, company, period))
    reg_rows = list(wb["Purchase Register"].iter_rows(values_only=True))
    reg_total = _find_row(reg_rows, col=0, value="TOTAL")
    assert reg_total[5] == 500
    assert reg_total[7] == 30
    party_row = next(r for r in reg_rows if r[2] == "Acme Supply")
    assert party_row[3] is None  # Supplier has no gst_number field


@pytest.mark.asyncio
async def test_payment_register_separates_receipts_and_payments(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    supplier = Supplier(company_id=company.id, name="Acme Supply")
    db.add_all([dealer, supplier])
    await db.flush()

    receivable_invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 1),
        due_date=date(2026, 7, 15),
        total=Decimal("1000"),
        status=InvoiceStatus.Partially_Paid,
    )
    payable_invoice = _invoice(
        company,
        direction=InvoiceDirection.payable,
        supplier=supplier,
        invoice_date=date(2026, 7, 1),
        due_date=date(2026, 7, 15),
        total=Decimal("400"),
        status=InvoiceStatus.Partially_Paid,
    )
    db.add_all([receivable_invoice, payable_invoice])
    await db.flush()
    db.add(
        _payment(company, receivable_invoice, amount=Decimal("600"), payment_date=date(2026, 7, 5))
    )
    db.add(
        _payment(company, payable_invoice, amount=Decimal("250"), payment_date=date(2026, 7, 6))
    )
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_payment_register_workbook(db, company, period))
    rows = list(wb["Payment Register"].iter_rows(values_only=True))

    receipts_row = _find_row(rows, col=1, value="Total Receipts")
    payments_row = _find_row(rows, col=1, value="Total Payments")
    net_row = _find_row(rows, col=1, value="Net (Receipts - Payments)")
    assert receipts_row[4] == 600
    assert payments_row[4] == 250
    assert net_row[4] == 350


@pytest.mark.asyncio
async def test_day_book_keeps_invoice_and_payment_totals_separate(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    db.add(dealer)
    await db.flush()

    invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 3),
        due_date=date(2026, 7, 18),
        total=Decimal("1000"),
        status=InvoiceStatus.Partially_Paid,
    )
    db.add(invoice)
    await db.flush()
    db.add(_payment(company, invoice, amount=Decimal("300"), payment_date=date(2026, 7, 4)))
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_day_book_workbook(db, company, period))
    rows = list(wb["Day Book"].iter_rows(values_only=True))

    # Never a single blended total mixing an accrual invoice amount with a
    # cash payment amount — see the "no blended financial metrics" principle.
    invoice_total_row = _find_row(rows, col=4, value="Total Invoices")
    payment_total_row = _find_row(rows, col=4, value="Total Payments")
    assert invoice_total_row[5] == 1000
    assert payment_total_row[5] == 300


# ── Aging ────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_aging_buckets_open_invoices_by_days_overdue(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    db.add(dealer)
    await db.flush()

    as_of = date(2026, 7, 19)
    invoices = [
        _invoice(  # not due yet
            company,
            direction=InvoiceDirection.receivable,
            dealer=dealer,
            invoice_date=as_of,
            due_date=as_of + timedelta(days=10),
            total=Decimal("100"),
        ),
        _invoice(  # 0-30 bucket
            company,
            direction=InvoiceDirection.receivable,
            dealer=dealer,
            invoice_date=as_of - timedelta(days=20),
            due_date=as_of - timedelta(days=10),
            total=Decimal("200"),
        ),
        _invoice(  # 31-60 bucket
            company,
            direction=InvoiceDirection.receivable,
            dealer=dealer,
            invoice_date=as_of - timedelta(days=60),
            due_date=as_of - timedelta(days=45),
            total=Decimal("300"),
        ),
        _invoice(  # 90+ bucket
            company,
            direction=InvoiceDirection.receivable,
            dealer=dealer,
            invoice_date=as_of - timedelta(days=140),
            due_date=as_of - timedelta(days=120),
            total=Decimal("400"),
        ),
    ]
    db.add_all(invoices)
    await db.commit()

    wb = _load(await aging.build_aging_report_workbook(db, company, as_of))
    rows = list(wb["Receivables Aging"].iter_rows(values_only=True))
    total_row = _find_row(rows, col=0, value="TOTAL")
    # headers: Party, Not Due, 0-30, 31-60, 61-90, 90+, Total
    assert total_row[1] == 100
    assert total_row[2] == 200
    assert total_row[3] == 300
    assert total_row[4] == 0
    assert total_row[5] == 400
    assert total_row[6] == 1000

    payables_rows = list(wb["Payables Aging"].iter_rows(values_only=True))
    assert not any(r[0] == "TOTAL" for r in payables_rows)  # nothing payable seeded


@pytest.mark.asyncio
async def test_aging_pdf_starts_with_pdf_magic_bytes(db: AsyncSession) -> None:
    company = await _make_company(db)
    out = await aging.build_aging_report_pdf(db, company, date(2026, 7, 19))
    assert out[:4] == b"%PDF"


# ── Registry ─────────────────────────────────────────────────────────────────


def test_registry_needs_party_only_for_ledger() -> None:
    assert {key for key, spec in REPORTS.items() if spec.needs_party} == {"ledger"}


def test_registry_pdf_only_for_ledger_and_aging() -> None:
    pdf_keys = {key for key, spec in REPORTS.items() if spec.build_pdf is not None}
    assert pdf_keys == {"ledger", "aging"}


@pytest.mark.asyncio
async def test_registry_full_report_matches_build_company_workbook(db: AsyncSession) -> None:
    company = await _make_company(db)
    out = await REPORTS["full"].build_xlsx(db, company, ReportContext())
    wb = _load(out)
    assert "Company" in wb.sheetnames
    assert "Invoices" in wb.sheetnames


@pytest.mark.asyncio
async def test_registry_pdf_builders_all_produce_valid_pdfs(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    db.add(dealer)
    await db.commit()
    await db.refresh(dealer)

    for spec in REPORTS.values():
        if spec.build_pdf is None:
            continue
        ctx = ReportContext(
            period=resolve_period(default_to_current_month=True),
            party=dealer if spec.needs_party else None,
            direction="receivable" if spec.needs_party else None,
        )
        out = await spec.build_pdf(db, company, ctx)
        assert out[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_registry_trend_report_has_three_sheets(db: AsyncSession) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Trend Registry Dealer")
    db.add(dealer)
    await db.commit()
    await db.refresh(dealer)
    # business_now, not date.today() — build_trend_report_workbook resolves
    # "today" via the company's business timezone (defaults to IST), and this
    # must land in the same calendar day for the seeded invoice to count.
    today = business_now(company.timezone).date()
    invoice = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=today,
        due_date=today + timedelta(days=14),
        total=Decimal("1000"),
    )
    db.add(invoice)
    await db.commit()

    out = await REPORTS["trend"].build_xlsx(db, company, ReportContext())
    wb = _load(out)
    assert wb.sheetnames == ["Cash Trend", "Dealer Trend", "Product Sales Trend"]
    cash_rows = list(wb["Cash Trend"].iter_rows(values_only=True))
    header_row = next(r for r in cash_rows if r[0] == "Metric")
    assert header_row == ("Metric", "This Week", "Last Week", "Δ")

    # Regression, same root cause as the ledger's freeze/filter test above:
    # every sheet in this workbook also calls write_meta_block before
    # write_header, so all three need prepare_sheet's fix, not just one.
    cash_ws = wb["Cash Trend"]
    assert cash_ws.freeze_panes == "A8"
    assert cash_ws.auto_filter.ref == "A7:D7"


@pytest.mark.asyncio
async def test_sales_register_includes_imported_invoices_with_no_line_items(
    db: AsyncSession,
) -> None:
    """An invoice with no InvoiceItem rows must still appear in the register.

    Regression: _register_rows inner-joined InvoiceItem, and only the WhatsApp
    order flow ever writes those — the CSV/Tally/Vyapar importers create the
    Invoice alone. So a distributor whose books arrived by Tally export saw a
    completely empty Sales/Purchase Register (header, no TOTAL row) while their
    Day Book and invoice lists showed the same invoices correctly.
    """
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Imported Dealer", gst_number="21ABCDE1234F1Z5")
    db.add(dealer)
    await db.flush()

    imported = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 12),
        due_date=date(2026, 7, 26),
        total=Decimal("11800.00"),
        status=InvoiceStatus.Pending,
    )
    # What the importer actually writes: real subtotal/GST split, zero items.
    imported.subtotal = Decimal("10000.00")
    imported.gst_amount = Decimal("1800.00")
    imported.source = InvoiceSource.csv_import
    db.add(imported)
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_sales_register_workbook(db, company, period))

    reg_rows = list(wb["Sales Register"].iter_rows(values_only=True))
    line = _find_row(reg_rows, col=1, value=imported.invoice_number)
    assert line[2] == "Imported Dealer"
    assert line[3] == "21ABCDE1234F1Z5"
    assert line[5] == 10000  # taxable value from the invoice's own subtotal
    assert line[6] == 18  # rate backed out of subtotal/gst_amount
    assert line[7] == 1800
    assert line[8] == 11800

    total = _find_row(reg_rows, col=0, value="TOTAL")
    assert total[5] == 10000
    assert total[7] == 1800

    # The rate-wise summary must agree with it, same invariant the itemised
    # register already guarantees.
    summary_rows = list(wb["Rate-wise Summary"].iter_rows(values_only=True))
    summary_data = summary_rows[1:-1]
    assert sum(r[1] for r in summary_data) == total[5]
    assert sum(r[2] for r in summary_data) == total[7]


@pytest.mark.asyncio
async def test_sales_register_mixes_itemised_and_imported_invoices_in_date_order(
    db: AsyncSession,
) -> None:
    company = await _make_company(db)
    dealer = Dealer(company_id=company.id, name="Mixed Dealer")
    db.add(dealer)
    await db.flush()

    imported = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 5),
        due_date=date(2026, 7, 20),
        total=Decimal("1050.00"),
    )
    imported.subtotal = Decimal("1000.00")
    imported.gst_amount = Decimal("50.00")
    itemised = _invoice(
        company,
        direction=InvoiceDirection.receivable,
        dealer=dealer,
        invoice_date=date(2026, 7, 20),
        due_date=date(2026, 8, 4),
        total=Decimal("224.00"),
    )
    db.add_all([imported, itemised])
    await db.flush()
    db.add(
        InvoiceItem(
            invoice_id=itemised.id,
            description="Soap",
            quantity=Decimal("2"),
            unit_price=Decimal("100"),
            line_total=Decimal("200"),
            gst_rate=Decimal("12"),
            gst_amount=Decimal("24"),
        )
    )
    await db.commit()

    period = resolve_period(month_str="2026-07")
    wb = _load(await registers.build_sales_register_workbook(db, company, period))
    reg_rows = list(wb["Sales Register"].iter_rows(values_only=True))

    data = [r for r in reg_rows if r[1] in (imported.invoice_number, itemised.invoice_number)]
    assert len(data) == 2
    # Sorted by invoice date across both sources, not itemised-then-imported.
    assert data[0][1] == imported.invoice_number
    assert data[1][1] == itemised.invoice_number

    total = _find_row(reg_rows, col=0, value="TOTAL")
    assert total[5] == 1200  # 1000 imported + 200 itemised
    assert total[7] == 74  # 50 + 24
