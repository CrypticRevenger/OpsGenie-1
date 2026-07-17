"""Per-company Excel data export tests — workbook builder, signed-link
generation/verification, and all three entry points (admin API, dashboard,
WhatsApp-triggered signed link).

    uv run alembic upgrade head
    uv run pytest tests/test_company_export.py -v
"""

from __future__ import annotations

import hashlib
import hmac
import io
import json
import uuid
from datetime import date
from decimal import Decimal

import pytest
from app.core.config import get_settings
from app.main import app
from app.models.company import Company, OnboardingState
from app.models.dealer import Dealer
from app.models.invoice import Invoice, InvoiceDirection, InvoiceSource, InvoiceStatus
from app.models.invoice_item import InvoiceItem
from app.models.payment import Payment, PaymentSource
from app.models.product import Product
from app.services.company_export import (
    build_company_workbook,
    generate_export_link,
    verify_export_link,
)
from app.services.whatsapp_client import WhatsAppSendResult
from httpx import ASGITransport, AsyncClient
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession


def _unique_phone() -> str:
    return f"+91{uuid.uuid4().int % 10_000_000_000:010d}"


async def _make_company(db: AsyncSession, name: str = "Export Test Co") -> Company:
    company = Company(business_name=name, owner_name="Owner", whatsapp_number=_unique_phone())
    db.add(company)
    await db.commit()
    await db.refresh(company)
    return company


async def _seed_data(db: AsyncSession, company: Company) -> None:
    dealer = Dealer(company_id=company.id, name="Ram Traders")
    db.add(dealer)
    await db.flush()
    product = Product(
        company_id=company.id,
        name="Rice",
        selling_price=Decimal("100.00"),
        purchase_price=Decimal("70.00"),
        stock_quantity=Decimal("50"),
    )
    db.add(product)
    await db.flush()
    today = date.today()
    invoice = Invoice(
        company_id=company.id,
        invoice_number=f"WA-{uuid.uuid4().hex[:10]}",
        direction=InvoiceDirection.receivable,
        dealer_id=dealer.id,
        invoice_date=today,
        due_date=today,
        subtotal=Decimal("1000.00"),
        gst_amount=Decimal("0.00"),
        total_amount=Decimal("1000.00"),
        status=InvoiceStatus.Partially_Paid,
        source=InvoiceSource.whatsapp,
    )
    db.add(invoice)
    await db.flush()
    db.add(
        InvoiceItem(
            invoice_id=invoice.id,
            product_id=product.id,
            description="Rice",
            quantity=Decimal("10"),
            unit_price=Decimal("100.00"),
            line_total=Decimal("1000.00"),
        )
    )
    db.add(
        Payment(
            company_id=company.id,
            invoice_id=invoice.id,
            amount=Decimal("400.00"),
            payment_date=today,
            source=PaymentSource.whatsapp,
        )
    )
    await db.commit()


def _load(wb_bytes: bytes):
    return load_workbook(io.BytesIO(wb_bytes))


# ── Workbook builder ─────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_fresh_company_produces_headers_only_sheets(db: AsyncSession) -> None:
    company = await _make_company(db)
    wb_bytes = await build_company_workbook(db, company)
    wb = _load(wb_bytes)

    expected_sheets = {
        "Company",
        "Products",
        "Dealers",
        "Suppliers",
        "Receivables",
        "Payables",
        "Invoices",
        "Payments",
        "Daily Business Summary",
        "FAQs",
    }
    assert expected_sheets.issubset(set(wb.sheetnames))
    assert "Activity Log" not in wb.sheetnames

    products_rows = list(wb["Products"].iter_rows(values_only=True))
    assert len(products_rows) == 1  # header row only


@pytest.mark.asyncio
async def test_workbook_reflects_real_data(db: AsyncSession) -> None:
    company = await _make_company(db)
    await _seed_data(db, company)

    wb = _load(await build_company_workbook(db, company))

    invoice_rows = list(wb["Invoices"].iter_rows(values_only=True))
    assert len(invoice_rows) == 3  # header + 1 invoice + total
    data_row = invoice_rows[1]
    assert data_row[0].startswith("WA-")
    assert data_row[2] == "Ram Traders"
    assert data_row[10] == 400  # Amount Paid
    assert data_row[11] == 600  # Amount Outstanding

    receivables_rows = list(wb["Receivables"].iter_rows(values_only=True))
    assert len(receivables_rows) == 3  # header + 1 invoice + total
    assert receivables_rows[1][4] == 600  # Amount Outstanding

    dealers_rows = list(wb["Dealers"].iter_rows(values_only=True))
    assert dealers_rows[1][0] == "Ram Traders"
    assert dealers_rows[1][-1] == 600  # Outstanding

    products_rows = list(wb["Products"].iter_rows(values_only=True))
    assert products_rows[1][0] == "Rice"
    assert products_rows[1][2] == 100  # selling_price
    assert products_rows[1][3] == 70  # purchase_price

    payments_rows = list(wb["Payments"].iter_rows(values_only=True))
    assert payments_rows[1][0] == 400


@pytest.mark.asyncio
async def test_company_sheet_has_export_metadata(db: AsyncSession) -> None:
    company = await _make_company(db, name="Metadata Co")
    wb = _load(await build_company_workbook(db, company))
    company_rows = list(wb["Company"].iter_rows(values_only=True))
    labels = [row[0] for row in company_rows]
    assert "Generated At" in labels
    assert "Generated By" in labels
    assert "Format Version" in labels
    assert ("Company", "Metadata Co") in company_rows


@pytest.mark.asyncio
async def test_export_never_leaks_across_companies(db: AsyncSession) -> None:
    company_a = await _make_company(db, name="Company A")
    company_b = await _make_company(db, name="Company B")
    await _seed_data(db, company_a)

    wb_a = _load(await build_company_workbook(db, company_a))
    wb_b = _load(await build_company_workbook(db, company_b))

    a_invoices = list(wb_a["Invoices"].iter_rows(values_only=True))
    b_invoices = list(wb_b["Invoices"].iter_rows(values_only=True))
    assert len(a_invoices) == 3  # header + 1 invoice + total
    assert len(b_invoices) == 1  # header only — no data leaked from company_a


# ── Signed link ──────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_signed_link_valid_then_tampered_then_expired(db: AsyncSession) -> None:
    company = await _make_company(db)
    link = generate_export_link(company, base_url="http://localhost:8000")
    _, cid, expires, sig = link.rsplit("/", 3)

    assert verify_export_link(uuid.UUID(cid), int(expires), sig) is True
    assert verify_export_link(uuid.UUID(cid), int(expires), "tampered") is False
    assert verify_export_link(uuid.UUID(cid), 1, sig) is False  # already expired


@pytest.mark.asyncio
async def test_signed_link_rejects_when_secret_unset(db: AsyncSession, monkeypatch) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "export_link_secret", None)
    company = await _make_company(db)
    with pytest.raises(ValueError):
        generate_export_link(company, base_url="http://localhost:8000")


# ── Admin API entry point ────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_admin_export_endpoint_returns_workbook(
    client: AsyncClient, db: AsyncSession
) -> None:
    company = await _make_company(db)
    await _seed_data(db, company)

    resp = await client.get(f"/admin/companies/{company.id}/export")
    assert resp.status_code == 200
    assert resp.headers["content-type"].startswith("application/vnd.openxmlformats")
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Invoices" in wb.sheetnames


@pytest.mark.asyncio
async def test_admin_export_endpoint_not_found(client: AsyncClient) -> None:
    resp = await client.get(f"/admin/companies/{uuid.uuid4()}/export")
    assert resp.status_code == 404


# ── Public signed endpoint ───────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_public_export_endpoint_valid_link(db: AsyncSession) -> None:
    company = await _make_company(db)
    link = generate_export_link(company, base_url="http://test")
    path = link.removeprefix("http://test")

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as anon_client:
        resp = await anon_client.get(path)
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Company" in wb.sheetnames


@pytest.mark.asyncio
async def test_public_export_endpoint_rejects_tampered_signature(db: AsyncSession) -> None:
    company = await _make_company(db)
    link = generate_export_link(company, base_url="http://test")
    path = link.removeprefix("http://test")[:-1] + "0"  # flip last signature char

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as anon_client:
        resp = await anon_client.get(path)
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_public_export_endpoint_rejects_expired_link(db: AsyncSession) -> None:
    company = await _make_company(db)
    settings = get_settings()
    signature = hmac.new(
        settings.export_link_secret.encode(), f"{company.id}:1".encode(), hashlib.sha256
    ).hexdigest()
    path = f"/export/{company.id}/1/{signature}"

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as anon_client:
        resp = await anon_client.get(path)
    assert resp.status_code == 403


# ── Dashboard entry point ────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_dashboard_export_requires_session(db: AsyncSession) -> None:
    company = await _make_company(db)
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as anon_client:
        resp = await anon_client.get(
            f"/dashboard/companies/{company.id}/export", follow_redirects=False
        )
    assert resp.status_code in (302, 303, 307)


@pytest.mark.asyncio
async def test_dashboard_export_authenticated_returns_workbook(db: AsyncSession) -> None:
    settings = get_settings()
    company = await _make_company(db)
    await _seed_data(db, company)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        await client.post("/dashboard/login", data={"password": settings.dashboard_password})
        resp = await client.get(f"/dashboard/companies/{company.id}/export")
    assert resp.status_code == 200
    wb = load_workbook(io.BytesIO(resp.content))
    assert "Invoices" in wb.sheetnames


# ── WhatsApp trigger ─────────────────────────────────────────────────────────


def _sign(body: bytes) -> str:
    secret = get_settings().whatsapp_app_secret
    digest = hmac.new(secret.encode(), body, hashlib.sha256).hexdigest()
    return f"sha256={digest}"


def _messages_payload(*, sender: str, text: str) -> dict:
    message = {
        "from": sender,
        "id": f"wamid.{uuid.uuid4().hex}",
        "timestamp": "1735689600",
        "type": "text",
        "text": {"body": text},
    }
    return {
        "object": "whatsapp_business_account",
        "entry": [
            {"id": "waba-id", "changes": [{"value": {"messages": [message]}, "field": "messages"}]}
        ],
    }


@pytest.mark.asyncio
async def test_whatsapp_export_trigger_replies_with_working_link(
    db: AsyncSession, monkeypatch
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "public_base_url", "http://test")

    phone = _unique_phone()
    bare_sender = phone.removeprefix("+")
    company = Company(
        business_name="Trigger Co",
        owner_name="Owner",
        whatsapp_number=phone,
        subscription_active=True,
        onboarding_state=OnboardingState.completed,
    )
    db.add(company)
    await db.commit()

    sent: list[str] = []

    async def _fake_send(to: str, body: str) -> WhatsAppSendResult:
        sent.append(body)
        return WhatsAppSendResult(message_id=f"wamid.{uuid.uuid4().hex}")

    monkeypatch.setattr("app.api.webhooks.whatsapp.send_text_message", _fake_send)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        body = json.dumps(_messages_payload(sender=bare_sender, text="export data")).encode()
        resp = await client.post(
            "/webhooks/whatsapp",
            content=body,
            headers={"Content-Type": "application/json", "X-Hub-Signature-256": _sign(body)},
        )
        assert resp.status_code == 200

        assert "Download" in sent[-1]
        link_line = next(line for line in sent[-1].splitlines() if line.startswith("Download"))
        url = link_line.split(": ", 1)[1]
        path = url.removeprefix("http://test")
        download_resp = await client.get(path)

    assert download_resp.status_code == 200


@pytest.mark.asyncio
async def test_whatsapp_export_trigger_without_config_says_so(
    db: AsyncSession, monkeypatch
) -> None:
    settings = get_settings()
    monkeypatch.setattr(settings, "public_base_url", None)

    phone = _unique_phone()
    bare_sender = phone.removeprefix("+")
    company = Company(
        business_name="No Config Co",
        owner_name="Owner",
        whatsapp_number=phone,
        subscription_active=True,
        onboarding_state=OnboardingState.completed,
    )
    db.add(company)
    await db.commit()

    sent: list[str] = []

    async def _fake_send(to: str, body: str) -> WhatsAppSendResult:
        sent.append(body)
        return WhatsAppSendResult(message_id=f"wamid.{uuid.uuid4().hex}")

    monkeypatch.setattr("app.api.webhooks.whatsapp.send_text_message", _fake_send)

    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        body = json.dumps(_messages_payload(sender=bare_sender, text="get my excel")).encode()
        resp = await client.post(
            "/webhooks/whatsapp",
            content=body,
            headers={"Content-Type": "application/json", "X-Hub-Signature-256": _sign(body)},
        )
        assert resp.status_code == 200

    assert "set up yet" in sent[-1].lower()
