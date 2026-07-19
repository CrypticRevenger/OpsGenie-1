"""Payment/receipt register, GST-annotated sales & purchase registers (with
a rate-wise summary sheet — the standard GSTR-1 register shape, taxable
value + GST split by rate, per invoice line), and the day book.

The sales/purchase register doubles as PROJECT_STATUS.md's "GST reports"
and "sales & purchase registers" bullets — one GST-annotated register per
direction satisfies both rather than building a near-duplicate report twice.
No CGST/SGST/IGST split: Dealer/Supplier carry no state/place-of-supply
field, so this is a register for manual GSTR-1 entry, not a filing-ready
export (out of scope, confirmed with the user).
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal

from openpyxl import Workbook
from sqlalchemy import Select, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.company import Company
from app.models.dealer import Dealer
from app.models.invoice import Invoice, InvoiceDirection
from app.models.invoice_item import InvoiceItem
from app.models.payment import Payment
from app.models.supplier import Supplier
from app.services.party_outstanding import Direction
from app.services.reports import xlsx_common
from app.services.reports.period import ReportPeriod
from app.services.reports.statuses import EXCLUDED_STATUSES

_REGISTER_HEADERS = (
    "Invoice Date",
    "Invoice No",
    "Party",
    "Party GSTIN",
    "Description",
    "Taxable Value",
    "GST Rate (%)",
    "GST Amount",
    "Line Total",
)
_REGISTER_FORMATS = ["date", None, None, None, None, "money", "qty", "money", "money"]

_SUMMARY_HEADERS = ("GST Rate (%)", "Taxable Value", "GST Amount", "Total Value")
_SUMMARY_FORMATS = ["qty", "money", "money", "money"]

_PAYMENT_REGISTER_HEADERS = ("Date", "Type", "Party", "Invoice Number", "Amount", "Method")
_PAYMENT_REGISTER_FORMATS = ["date", None, None, None, "money", None]

_DAY_BOOK_HEADERS = ("Date", "Type", "Direction", "Party", "Reference", "Amount")
_DAY_BOOK_FORMATS = ["date", None, None, None, None, "money"]


async def _names_by_id(db: AsyncSession, model: type[Dealer] | type[Supplier], ids: set) -> dict:
    """id -> name for a batch of Dealer/Supplier rows — the payment register
    and day book both need this twice (dealer side + supplier side).
    """
    if not ids:
        return {}
    rows = (await db.scalars(select(model).where(model.id.in_(ids)))).all()
    return {row.id: row.name for row in rows}


def _payment_rows_stmt(company_id: uuid.UUID) -> Select:
    return (
        select(
            Payment,
            Invoice.invoice_number,
            Invoice.direction,
            Invoice.dealer_id,
            Invoice.supplier_id,
        )
        .join(Invoice, Payment.invoice_id == Invoice.id)
        .where(Payment.company_id == company_id)
    )


async def _register_rows(
    db: AsyncSession, company_id: uuid.UUID, direction: Direction, period: ReportPeriod
):
    party_model = Dealer if direction == "receivable" else Supplier
    party_column = Invoice.dealer_id if direction == "receivable" else Invoice.supplier_id
    stmt = (
        select(InvoiceItem, Invoice.invoice_number, Invoice.invoice_date, party_column)
        .join(Invoice, InvoiceItem.invoice_id == Invoice.id)
        .where(
            Invoice.company_id == company_id,
            Invoice.direction == InvoiceDirection(direction),
            Invoice.status.notin_(EXCLUDED_STATUSES),
        )
        .order_by(Invoice.invoice_date, Invoice.invoice_number)
    )
    if period.from_date is not None:
        stmt = stmt.where(Invoice.invoice_date >= period.from_date)
    if period.to_date is not None:
        stmt = stmt.where(Invoice.invoice_date <= period.to_date)
    rows = (await db.execute(stmt)).all()

    party_ids = {r[3] for r in rows if r[3] is not None}
    parties = {}
    if party_ids:
        matched = await db.scalars(select(party_model).where(party_model.id.in_(party_ids)))
        parties = {p.id: p for p in matched.all()}
    return rows, parties


async def _build_register_workbook(
    db: AsyncSession,
    company: Company,
    period: ReportPeriod,
    *,
    direction: Direction,
    sheet_title: str,
) -> bytes:
    rows, parties = await _register_rows(db, company.id, direction, period)
    generated_at = datetime.now(UTC)

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(sheet_title)
    xlsx_common.write_meta_block(
        ws,
        report_name=sheet_title,
        company_name=company.business_name,
        period_label=period.label,
        generated_at=generated_at,
        row_count=len(rows),
    )
    xlsx_common.write_header(ws, list(_REGISTER_HEADERS))

    rate_totals: dict[Decimal, list[Decimal]] = {}
    total_taxable = Decimal("0")
    total_gst = Decimal("0")
    total_value = Decimal("0")
    for item, invoice_number, invoice_date, party_id in rows:
        party = parties.get(party_id)
        party_name = party.name if party is not None else "Unknown"
        party_gstin = getattr(party, "gst_number", None) if party is not None else None
        line_total_incl = item.line_total + item.gst_amount
        xlsx_common.row(
            ws,
            [
                invoice_date,
                invoice_number,
                party_name,
                party_gstin,
                item.description,
                item.line_total,
                item.gst_rate,
                item.gst_amount,
                line_total_incl,
            ],
            _REGISTER_FORMATS,
        )
        bucket = rate_totals.setdefault(item.gst_rate, [Decimal("0"), Decimal("0"), Decimal("0")])
        bucket[0] += item.line_total
        bucket[1] += item.gst_amount
        bucket[2] += line_total_incl
        total_taxable += item.line_total
        total_gst += item.gst_amount
        total_value += line_total_incl
    if rows:
        xlsx_common.total_row(
            ws,
            ["TOTAL", None, None, None, None, total_taxable, None, total_gst, total_value],
            _REGISTER_FORMATS,
        )

    summary_ws = wb.create_sheet("Rate-wise Summary")
    xlsx_common.write_header(summary_ws, list(_SUMMARY_HEADERS))
    for rate in sorted(rate_totals):
        taxable, gst, total = rate_totals[rate]
        xlsx_common.row(summary_ws, [rate, taxable, gst, total], _SUMMARY_FORMATS)
    if rate_totals:
        xlsx_common.total_row(
            summary_ws, ["TOTAL", total_taxable, total_gst, total_value], _SUMMARY_FORMATS
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def build_sales_register_workbook(
    db: AsyncSession, company: Company, period: ReportPeriod
) -> bytes:
    return await _build_register_workbook(
        db, company, period, direction="receivable", sheet_title="Sales Register"
    )


async def build_purchase_register_workbook(
    db: AsyncSession, company: Company, period: ReportPeriod
) -> bytes:
    return await _build_register_workbook(
        db, company, period, direction="payable", sheet_title="Purchase Register"
    )


async def build_payment_register_workbook(
    db: AsyncSession, company: Company, period: ReportPeriod
) -> bytes:
    stmt = _payment_rows_stmt(company.id).order_by(Payment.payment_date, Payment.created_at)
    if period.from_date is not None:
        stmt = stmt.where(Payment.payment_date >= period.from_date)
    if period.to_date is not None:
        stmt = stmt.where(Payment.payment_date <= period.to_date)
    rows = (await db.execute(stmt)).all()

    dealer_names = await _names_by_id(db, Dealer, {r[3] for r in rows if r[3] is not None})
    supplier_names = await _names_by_id(db, Supplier, {r[4] for r in rows if r[4] is not None})

    generated_at = datetime.now(UTC)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Payment Register")
    xlsx_common.write_meta_block(
        ws,
        report_name="Payment Register",
        company_name=company.business_name,
        period_label=period.label,
        generated_at=generated_at,
        row_count=len(rows),
    )
    xlsx_common.write_header(ws, list(_PAYMENT_REGISTER_HEADERS))

    total_receipts = Decimal("0")
    total_payments = Decimal("0")
    for payment, invoice_number, direction, dealer_id, supplier_id in rows:
        is_receipt = direction == InvoiceDirection.receivable
        party_name = (
            dealer_names.get(dealer_id, "Unknown")
            if is_receipt
            else supplier_names.get(supplier_id, "Unknown")
        )
        xlsx_common.row(
            ws,
            [
                payment.payment_date,
                "Receipt" if is_receipt else "Payment",
                party_name,
                invoice_number,
                payment.amount,
                payment.method,
            ],
            _PAYMENT_REGISTER_FORMATS,
        )
        if is_receipt:
            total_receipts += payment.amount
        else:
            total_payments += payment.amount
    if rows:
        xlsx_common.total_row(
            ws, ["", "Total Receipts", "", "", total_receipts, None], _PAYMENT_REGISTER_FORMATS
        )
        xlsx_common.total_row(
            ws, ["", "Total Payments", "", "", total_payments, None], _PAYMENT_REGISTER_FORMATS
        )
        # Receipts/payments are both already cash-basis, so netting them here
        # is a real cash-movement figure (same as daily_business_snapshot's
        # net_cash_movement) — not a blend of accrual and cash, which is what
        # this codebase otherwise never does.
        xlsx_common.total_row(
            ws,
            ["", "Net (Receipts - Payments)", "", "", total_receipts - total_payments, None],
            _PAYMENT_REGISTER_FORMATS,
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def build_day_book_workbook(
    db: AsyncSession, company: Company, period: ReportPeriod
) -> bytes:
    invoice_stmt = select(Invoice).where(
        Invoice.company_id == company.id, Invoice.status.notin_(EXCLUDED_STATUSES)
    )
    if period.from_date is not None:
        invoice_stmt = invoice_stmt.where(Invoice.invoice_date >= period.from_date)
    if period.to_date is not None:
        invoice_stmt = invoice_stmt.where(Invoice.invoice_date <= period.to_date)
    invoices = (await db.scalars(invoice_stmt)).all()

    payment_stmt = _payment_rows_stmt(company.id)
    if period.from_date is not None:
        payment_stmt = payment_stmt.where(Payment.payment_date >= period.from_date)
    if period.to_date is not None:
        payment_stmt = payment_stmt.where(Payment.payment_date <= period.to_date)
    payment_rows = (await db.execute(payment_stmt)).all()

    dealer_ids = {inv.dealer_id for inv in invoices if inv.dealer_id} | {
        r[3] for r in payment_rows if r[3] is not None
    }
    supplier_ids = {inv.supplier_id for inv in invoices if inv.supplier_id} | {
        r[4] for r in payment_rows if r[4] is not None
    }
    dealer_names = await _names_by_id(db, Dealer, dealer_ids)
    supplier_names = await _names_by_id(db, Supplier, supplier_ids)

    entries: list[tuple] = []
    for inv in invoices:
        is_receivable = inv.direction == InvoiceDirection.receivable
        party_name = (
            dealer_names.get(inv.dealer_id, "Unknown")
            if is_receivable
            else supplier_names.get(inv.supplier_id, "Unknown")
        )
        entries.append(
            (
                inv.invoice_date,
                "Invoice",
                "Receivable" if is_receivable else "Payable",
                party_name,
                inv.invoice_number,
                inv.total_amount,
            )
        )
    for payment, invoice_number, direction, dealer_id, supplier_id in payment_rows:
        is_receivable = direction == InvoiceDirection.receivable
        party_name = (
            dealer_names.get(dealer_id, "Unknown")
            if is_receivable
            else supplier_names.get(supplier_id, "Unknown")
        )
        entries.append(
            (
                payment.payment_date,
                "Payment",
                "Receivable" if is_receivable else "Payable",
                party_name,
                invoice_number,
                payment.amount,
            )
        )
    entries.sort(key=lambda e: e[0])

    generated_at = datetime.now(UTC)
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Day Book")
    xlsx_common.write_meta_block(
        ws,
        report_name="Day Book",
        company_name=company.business_name,
        period_label=period.label,
        generated_at=generated_at,
        row_count=len(entries),
    )
    xlsx_common.write_header(ws, list(_DAY_BOOK_HEADERS))

    # Invoice totals (accrual) and payment totals (cash) are summed
    # separately, never blended into one figure — see "no blended financial
    # metrics" in this project's working principles.
    total_invoices = Decimal("0")
    total_payments = Decimal("0")
    for entry_date, kind, direction_label, party_name, reference, amount in entries:
        xlsx_common.row(
            ws,
            [entry_date, kind, direction_label, party_name, reference, amount],
            _DAY_BOOK_FORMATS,
        )
        if kind == "Invoice":
            total_invoices += amount
        else:
            total_payments += amount
    if entries:
        xlsx_common.total_row(
            ws, ["", "", "", "", "Total Invoices", total_invoices], _DAY_BOOK_FORMATS
        )
        xlsx_common.total_row(
            ws, ["", "", "", "", "Total Payments", total_payments], _DAY_BOOK_FORMATS
        )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
