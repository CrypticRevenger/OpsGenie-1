"""Shared openpyxl styling primitives for every generated workbook.

Extracted verbatim from app/services/company_export.py (the original
all-time workbook) so the new period-scoped reports in this package render
with the exact same look — bold/frozen/filterable headers, sized columns,
₹/date number formats — instead of a second, drifting copy of the same
styling code. company_export.py imports these back rather than redefining
them, so its own output is byte-for-byte unchanged.
"""

from __future__ import annotations

from datetime import datetime

from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
TOTAL_FONT = Font(bold=True)
TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

MONEY_FMT = "₹#,##0.00"
QTY_FMT = "#,##0.##"
DATE_FMT = "dd-mmm-yyyy"
NUMBER_FORMATS = {"money": MONEY_FMT, "qty": QTY_FMT, "date": DATE_FMT}

# Columns that need to be wider than the default heuristic to hold their
# typical content without truncating in the sheet view.
WIDE_COLUMNS = {"Details", "Notes", "Address", "Invoice Number", "Business Name"}

# write_meta_block always writes exactly this many rows (5 label/value pairs
# + 1 blank separator) before returning control to its caller — a fixed,
# input-independent count, so a caller that needs to know where the meta
# block will end (to call prepare_sheet beforehand — see its docstring)
# doesn't need to run it first just to find out.
META_BLOCK_ROWS = 6
META_BLOCK_HEADER_ROW = META_BLOCK_ROWS + 1


def s(value: object) -> str | None:
    """Render an enum/UUID/None as a plain string cell value."""
    if value is None:
        return None
    return str(value)


def autosize(ws: Worksheet, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        width = 34 if header in WIDE_COLUMNS else max(len(header) + 6, 14)
        ws.column_dimensions[letter].width = width


def prepare_sheet(ws: Worksheet, headers: list[str], *, start_row: int = 1) -> None:
    """Freeze panes, auto-filter, and column widths — verified empirically
    (openpyxl's write-only writer serializes each row's worth of sheet state
    as soon as it's appended) that write-only mode only honors any of these
    three when set before the sheet's FIRST ws.append() of any kind; set
    afterward, they silently no-op and the reload has none of them. This
    must therefore run before write_meta_block too, not just before
    write_header — previously write_header tried to set all three itself,
    which worked for a sheet with nothing above the header (company_export.py's
    sheets, which don't need this function at all) but silently did nothing
    on every period-scoped report in this package, all of which call
    write_meta_block — a header row is always freeze/filter/width row 1 there
    on reload, and the real header, several rows lower, has neither.

    `start_row` is the 1-indexed row the real column header will occupy — 1
    for a sheet with nothing above it, or META_BLOCK_HEADER_ROW for one that
    calls write_meta_block next.
    """
    ws.freeze_panes = f"A{start_row + 1}"
    ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row}"
    autosize(ws, headers)


def write_header(ws: Worksheet, headers: list[str], *, start_row: int = 1) -> None:
    """Bold white-on-blue header row — every sheet in every workbook uses
    this same look. Also calls prepare_sheet for the common case
    (start_row=1, i.e. nothing above the header on this sheet), matching
    this function's original all-in-one behavior; a sheet that calls
    write_meta_block first must call prepare_sheet(..., start_row=
    META_BLOCK_HEADER_ROW) itself, before write_meta_block — by the time
    write_header runs on that sheet it's too late (see prepare_sheet).
    """
    if start_row == 1:
        prepare_sheet(ws, headers, start_row=start_row)
    cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cells.append(cell)
    ws.append(cells)


def row(ws: Worksheet, values: list, formats: list[str | None]) -> None:
    """Append a data row, wrapping only the cells that need a number format —
    plain values pass through untouched.
    """
    cells = []
    for value, fmt in zip(values, formats, strict=True):
        if fmt is None or value is None:
            cells.append(value)
            continue
        cell = WriteOnlyCell(ws, value=value)
        cell.number_format = NUMBER_FORMATS[fmt]
        cells.append(cell)
    ws.append(cells)


def total_row(ws: Worksheet, values: list, formats: list[str | None]) -> None:
    """Bold, shaded total row — values/formats align positionally with the
    sheet's header, so callers pass None for any column that isn't summed.
    """
    cells = []
    for value, fmt in zip(values, formats, strict=True):
        cell = WriteOnlyCell(ws, value=value)
        cell.font = TOTAL_FONT
        cell.fill = TOTAL_FILL
        if fmt is not None and value is not None:
            cell.number_format = NUMBER_FORMATS[fmt]
        cells.append(cell)
    ws.append(cells)


def write_meta_block(
    ws: Worksheet,
    *,
    report_name: str,
    company_name: str,
    period_label: str,
    generated_at: datetime,
    row_count: int,
) -> int:
    """A small label/value block at the top of a report's primary sheet —
    which report, which company, which period, how many rows, generated
    when — so opening the file (or forwarding it) never leaves the reader
    guessing what they're looking at. Same two-column style
    company_export.py's own Company sheet already uses.

    Returns the 1-indexed row the real header should start at — pass it as
    write_header(..., start_row=...) so freeze_panes/auto_filter anchor to
    the actual header row, not wherever this block happens to end.
    """
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 44
    meta_font = Font(bold=True)
    rows_written = 0
    for label, value in [
        ("Report", report_name),
        ("Company", company_name),
        ("Period", period_label),
        ("Rows", row_count),
        ("Generated At", generated_at.isoformat()),
    ]:
        label_cell = WriteOnlyCell(ws, value=label)
        label_cell.font = meta_font
        ws.append([label_cell, value])
        rows_written += 1
    ws.append([])
    rows_written += 1
    return rows_written + 1
