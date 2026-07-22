"""Outstanding (aging) report — every open invoice bucketed by days overdue
(Not Due / 0-30 / 31-60 / 61-90 / 90+), grouped per party, the standard
Tally-style aging matrix. A snapshot as of a given date, not a from/to
period. Uses party_outstanding.py's current-status-based
calculate_outstanding_for_company* (Invoice.status.in_(OPEN_STATUSES)), not
its date-bounded calculate_outstanding_for_company_as_of sibling (added for
trend_analytics.py's week/month-over-month comparisons) — an aging bucket is
inherently about invoices open *right now*, so `as_of` is always effectively
"now" in practice here regardless of which helper is used.
"""

from __future__ import annotations

import io
import uuid
from datetime import UTC, date, datetime
from decimal import Decimal

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.company import Company
from app.models.dealer import Dealer
from app.models.invoice import Invoice, InvoiceDirection
from app.models.payment import Payment
from app.models.supplier import Supplier
from app.services.money_format import format_inr
from app.services.party_outstanding import Direction
from app.services.reports import pdf_common, xlsx_common
from app.services.reports.statuses import OPEN_STATUSES

_AGING_BUCKETS = ("Not Due", "0-30", "31-60", "61-90", "90+")
_AGING_HEADERS = ("Party", *_AGING_BUCKETS, "Total")
_AGING_FORMATS = [None, "money", "money", "money", "money", "money", "money"]
_PDF_COL_WIDTHS = (40, 24, 24, 24, 24, 24, 28)


def _bucket_for(days_overdue: int) -> str:
    if days_overdue <= 0:
        return "Not Due"
    if days_overdue <= 30:
        return "0-30"
    if days_overdue <= 60:
        return "31-60"
    if days_overdue <= 90:
        return "61-90"
    return "90+"


async def _aging_by_party(
    db: AsyncSession, company_id: uuid.UUID, direction: Direction, as_of: date
) -> tuple[dict[uuid.UUID, dict[str, Decimal]], dict[uuid.UUID, Dealer | Supplier]]:
    party_model = Dealer if direction == "receivable" else Supplier
    invoices = (
        await db.scalars(
            select(Invoice).where(
                Invoice.company_id == company_id,
                Invoice.direction == InvoiceDirection(direction),
                Invoice.status.in_(OPEN_STATUSES),
            )
        )
    ).all()
    if not invoices:
        return {}, {}

    invoice_ids = [inv.id for inv in invoices]
    paid_rows = await db.execute(
        select(Payment.invoice_id, func.sum(Payment.amount))
        .where(Payment.invoice_id.in_(invoice_ids))
        .group_by(Payment.invoice_id)
    )
    paid_by_invoice: dict[uuid.UUID, Decimal] = dict(paid_rows.all())

    party_ids = {
        (inv.dealer_id if direction == "receivable" else inv.supplier_id) for inv in invoices
    }
    party_ids.discard(None)
    matched = await db.scalars(select(party_model).where(party_model.id.in_(party_ids)))
    parties = {p.id: p for p in matched.all()}

    buckets: dict[uuid.UUID, dict[str, Decimal]] = {}
    for inv in invoices:
        party_id = inv.dealer_id if direction == "receivable" else inv.supplier_id
        if party_id is None:
            continue
        paid = paid_by_invoice.get(inv.id, Decimal("0"))
        outstanding = inv.total_amount - paid
        if outstanding <= 0:
            continue
        bucket = _bucket_for((as_of - inv.due_date).days)
        party_buckets = buckets.setdefault(party_id, {b: Decimal("0") for b in _AGING_BUCKETS})
        party_buckets[bucket] += outstanding
    return buckets, parties


def _sorted_party_ids(buckets: dict, parties: dict) -> list:
    return sorted(buckets, key=lambda pid: parties[pid].name if pid in parties else "")


def _write_aging_sheet(
    ws: Worksheet,
    buckets: dict[uuid.UUID, dict[str, Decimal]],
    parties: dict,
    *,
    company_name: str,
    period_label: str,
    generated_at: datetime,
    sheet_label: str,
) -> None:
    xlsx_common.write_meta_block(
        ws,
        report_name=sheet_label,
        company_name=company_name,
        period_label=period_label,
        generated_at=generated_at,
        row_count=len(buckets),
    )
    xlsx_common.write_header(ws, list(_AGING_HEADERS))

    grand_totals = {b: Decimal("0") for b in _AGING_BUCKETS}
    grand_total = Decimal("0")
    for party_id in _sorted_party_ids(buckets, parties):
        party = parties.get(party_id)
        name = party.name if party is not None else "Unknown"
        row_buckets = buckets[party_id]
        total = sum(row_buckets.values(), Decimal("0"))
        xlsx_common.row(
            ws, [name, *[row_buckets[b] for b in _AGING_BUCKETS], total], _AGING_FORMATS
        )
        for b in _AGING_BUCKETS:
            grand_totals[b] += row_buckets[b]
        grand_total += total
    if buckets:
        xlsx_common.total_row(
            ws, ["TOTAL", *[grand_totals[b] for b in _AGING_BUCKETS], grand_total], _AGING_FORMATS
        )


async def build_aging_report_workbook(db: AsyncSession, company: Company, as_of: date) -> bytes:
    receivable_buckets, dealers = await _aging_by_party(db, company.id, "receivable", as_of)
    payable_buckets, suppliers = await _aging_by_party(db, company.id, "payable", as_of)
    generated_at = datetime.now(UTC)
    period_label = f"As of {as_of.isoformat()}"

    wb = Workbook(write_only=True)
    _write_aging_sheet(
        wb.create_sheet("Receivables Aging"),
        receivable_buckets,
        dealers,
        company_name=company.business_name,
        period_label=period_label,
        generated_at=generated_at,
        sheet_label="Receivables Aging",
    )
    _write_aging_sheet(
        wb.create_sheet("Payables Aging"),
        payable_buckets,
        suppliers,
        company_name=company.business_name,
        period_label=period_label,
        generated_at=generated_at,
        sheet_label="Payables Aging",
    )

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


async def build_aging_report_pdf(db: AsyncSession, company: Company, as_of: date) -> bytes:
    receivable_buckets, dealers = await _aging_by_party(db, company.id, "receivable", as_of)
    payable_buckets, suppliers = await _aging_by_party(db, company.id, "payable", as_of)
    generated_at = datetime.now(UTC)
    period_label = f"As of {as_of.isoformat()}"

    pdf = FPDF(format="A4")
    pdf.add_page()
    unicode_ok = pdf_common.try_load_unicode_fonts(pdf)

    def money(amount: Decimal) -> str:
        return format_inr(amount) if unicode_ok else pdf_common.rs(amount)

    def render_section(buckets: dict, parties: dict, section_title: str) -> None:
        pdf_common.write_meta_header(
            pdf,
            report_name=section_title,
            company_name=company.business_name,
            period_label=period_label,
            generated_at=generated_at,
            unicode_ok=unicode_ok,
        )
        rows: list[tuple[object, ...]] = []
        grand_totals = {b: Decimal("0") for b in _AGING_BUCKETS}
        grand_total = Decimal("0")
        for party_id in _sorted_party_ids(buckets, parties):
            party = parties.get(party_id)
            name = party.name if party is not None else "Unknown"
            row_buckets = buckets[party_id]
            total = sum(row_buckets.values(), Decimal("0"))
            rows.append((name, *[money(row_buckets[b]) for b in _AGING_BUCKETS], money(total)))
            for b in _AGING_BUCKETS:
                grand_totals[b] += row_buckets[b]
            grand_total += total
        if rows:
            rows.append(
                ("TOTAL", *[money(grand_totals[b]) for b in _AGING_BUCKETS], money(grand_total))
            )
        pdf_common.write_table(
            pdf,
            headers=_AGING_HEADERS,
            col_widths=_PDF_COL_WIDTHS,
            rows=rows,
            unicode_ok=unicode_ok,
            bold_rows={len(rows) - 1} if rows else set(),
        )

    render_section(receivable_buckets, dealers, "Receivables Aging")
    pdf.add_page()
    render_section(payable_buckets, suppliers, "Payables Aging")

    return bytes(pdf.output())
