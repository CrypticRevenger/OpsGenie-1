"""Business Trends report — week-over-week cash, 30-day dealer and product
movement, exported as one workbook. Always "as of today" (ctx.period is
ignored), same convention app/services/reports/aging.py already uses for its
own always-current report — the trend windows aren't caller-selectable.
"""

from __future__ import annotations

import io
from datetime import UTC, datetime

from openpyxl import Workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.company import Company
from app.services.reports import xlsx_common
from app.services.snapshot import business_now
from app.services.trend_analytics import DealerTrend, ProductSalesTrend, build_trend_report

_CASH_FORMATS = [None, "money", "money", "money"]
_DEALER_HEADERS = (
    "Dealer",
    "Orders (this 30d)",
    "Orders (prior 30d)",
    "Order Δ",
    "Value (this 30d)",
    "Value (prior 30d)",
    "Value Δ",
    "Outstanding (now)",
    "Outstanding (30d ago)",
    "Outstanding Δ",
)
_DEALER_FORMATS = [None, None, None, None, "money", "money", "money", "money", "money", "money"]
_PRODUCT_HEADERS = (
    "Product",
    "Units (this 30d)",
    "Units (prior 30d)",
    "Unit Δ",
    "Revenue (this 30d)",
    "Revenue (prior 30d)",
    "Revenue Δ",
)
_PRODUCT_FORMATS = [None, "qty", "qty", "qty", "money", "money", "money"]


def _dealer_row(t: DealerTrend) -> list:
    return [
        t.dealer_name,
        t.orders_current,
        t.orders_prior,
        t.orders_current - t.orders_prior,
        t.value_current,
        t.value_prior,
        t.value_delta,
        t.outstanding_current,
        t.outstanding_prior,
        t.outstanding_delta,
    ]


def _product_row(t: ProductSalesTrend) -> list:
    return [
        t.product_name,
        t.units_current,
        t.units_prior,
        t.units_delta,
        t.revenue_current,
        t.revenue_prior,
        t.revenue_current - t.revenue_prior,
    ]


async def build_trend_report_workbook(db: AsyncSession, company: Company) -> bytes:
    today = business_now(company.timezone).date()
    report = await build_trend_report(db, company.id, today)
    generated_at = datetime.now(UTC)

    wb = Workbook(write_only=True)

    cash_ws = wb.create_sheet("Cash Trend")
    cash_period = (
        f"This week ({report.cash_window.current_start.isoformat()} to "
        f"{report.cash_window.current_end.isoformat()}) vs prior week "
        f"({report.cash_window.prior_start.isoformat()} to "
        f"{report.cash_window.prior_end.isoformat()})"
    )
    xlsx_common.write_meta_block(
        cash_ws,
        report_name="Cash Trend",
        company_name=company.business_name,
        period_label=cash_period,
        generated_at=generated_at,
        row_count=4,
    )
    xlsx_common.write_header(cash_ws, ["Metric", "This Week", "Last Week", "Δ"])
    cash = report.cash
    xlsx_common.row(
        cash_ws,
        [
            "Collections",
            cash.collections_current,
            cash.collections_prior,
            cash.collections_current - cash.collections_prior,
        ],
        _CASH_FORMATS,
    )
    xlsx_common.row(
        cash_ws,
        [
            "Supplier Payments",
            cash.payments_current,
            cash.payments_prior,
            cash.payments_current - cash.payments_prior,
        ],
        _CASH_FORMATS,
    )
    xlsx_common.total_row(
        cash_ws,
        [
            "Net Cash Movement",
            cash.net_cash_current,
            cash.net_cash_prior,
            cash.net_cash_current - cash.net_cash_prior,
        ],
        _CASH_FORMATS,
    )
    xlsx_common.row(
        cash_ws,
        [
            "Sales (Invoiced)",
            cash.sales_current,
            cash.sales_prior,
            cash.sales_current - cash.sales_prior,
        ],
        _CASH_FORMATS,
    )

    dealer_ws = wb.create_sheet("Dealer Trend")
    activity_period = (
        f"This 30 days ({report.activity_window.current_start.isoformat()} to "
        f"{report.activity_window.current_end.isoformat()}) vs prior 30 days "
        f"({report.activity_window.prior_start.isoformat()} to "
        f"{report.activity_window.prior_end.isoformat()})"
    )
    xlsx_common.write_meta_block(
        dealer_ws,
        report_name="Dealer Trend",
        company_name=company.business_name,
        period_label=activity_period,
        generated_at=generated_at,
        row_count=len(report.dealers),
    )
    xlsx_common.write_header(dealer_ws, list(_DEALER_HEADERS))
    for trend in sorted(report.dealers, key=lambda t: t.outstanding_delta, reverse=True):
        xlsx_common.row(dealer_ws, _dealer_row(trend), _DEALER_FORMATS)

    product_ws = wb.create_sheet("Product Sales Trend")
    xlsx_common.write_meta_block(
        product_ws,
        report_name="Product Sales Trend",
        company_name=company.business_name,
        period_label=activity_period,
        generated_at=generated_at,
        row_count=len(report.products),
    )
    xlsx_common.write_header(product_ws, list(_PRODUCT_HEADERS))
    for trend in sorted(report.products, key=lambda t: t.units_delta, reverse=True):
        xlsx_common.row(product_ws, _product_row(trend), _PRODUCT_FORMATS)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
