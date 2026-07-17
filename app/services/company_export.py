"""Per-company Excel data export.

Two independent pieces:
- build_company_workbook: a pure read, rebuilt fresh from the live database
  on every call. Deliberately NOT a stored/incrementally-updated file — this
  codebase has dozens of independent write paths (guided WhatsApp workflows,
  admin CRUD, CSV import, onboarding); hooking every one of them to also
  patch a persisted .xlsx would be exactly the kind of easy-to-miss-a-spot
  design that causes silent staleness. A full rebuild can never be "half
  updated." Uses openpyxl's write_only streaming writer (already a project
  dependency) plus server-side streamed reads for the large tables, so
  memory stays bounded regardless of how many rows a company accumulates.
  Every sheet gets a styled header (bold/frozen/filterable), sized columns,
  and ₹/date number formats so the file reads as a finished report rather
  than a raw data dump.
- generate_export_link: a short-lived HMAC-signed URL for the WhatsApp-facing
  download path, verified statelessly by app/api/export.py — nothing is
  written to the database to "issue" a link, so there's nothing to store or
  revoke. Same signing pattern this codebase already uses to verify Meta's
  own webhook payloads (app/api/webhooks/whatsapp.py's X-Hub-Signature-256).
"""

from __future__ import annotations

import hashlib
import hmac
import io
import time
import uuid
from datetime import UTC, datetime
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import get_settings
from app.models.company import Company
from app.models.daily_business_snapshot import DailyBusinessSnapshot
from app.models.dealer import Dealer
from app.models.faq import FAQ
from app.models.invoice import Invoice, InvoiceDirection, InvoiceStatus
from app.models.payment import Payment
from app.models.product import Product
from app.models.supplier import Supplier
from app.services.party_outstanding import calculate_outstanding_for_company

EXPORT_FORMAT_VERSION = "1.1"

_OPEN_STATUSES = (InvoiceStatus.Pending, InvoiceStatus.Partially_Paid, InvoiceStatus.Overdue)
_STREAM_CHUNK = 500

# ── Sheet styling ────────────────────────────────────────────────────────────
# write_only Workbooks can't be styled with the usual ws['A1'].font = ... —
# every styled cell has to be built as a WriteOnlyCell before it's appended.

_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_TOTAL_FONT = Font(bold=True)
_TOTAL_FILL = PatternFill("solid", fgColor="D9E1F2")

_MONEY_FMT = "₹#,##0.00"
_QTY_FMT = "#,##0.##"
_DATE_FMT = "dd-mmm-yyyy"
_NUMBER_FORMATS = {"money": _MONEY_FMT, "qty": _QTY_FMT, "date": _DATE_FMT}

# Columns that need to be wider than the default heuristic to hold their
# typical content without truncating in the sheet view.
_WIDE_COLUMNS = {"Details", "Notes", "Address", "Invoice Number", "Business Name"}


def _s(value: object) -> str | None:
    """Render an enum/UUID/None as a plain string cell value."""
    if value is None:
        return None
    return str(value)


def _autosize(ws: Worksheet, headers: list[str]) -> None:
    for idx, header in enumerate(headers, start=1):
        letter = get_column_letter(idx)
        width = 34 if header in _WIDE_COLUMNS else max(len(header) + 6, 14)
        ws.column_dimensions[letter].width = width


def _write_header(ws: Worksheet, headers: list[str]) -> None:
    """Bold white-on-blue header row, frozen and filterable, plus sized
    columns — every sheet in the workbook uses this same look. In write_only
    mode, sheet-level properties like freeze_panes only take effect if set
    before the first append(), so this must run before the header row is
    written.
    """
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    _autosize(ws, headers)
    cells = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cells.append(cell)
    ws.append(cells)


def _row(ws: Worksheet, values: list, formats: list[str | None]) -> None:
    """Append a data row, wrapping only the cells that need a number format —
    plain values pass through untouched.
    """
    cells = []
    for value, fmt in zip(values, formats, strict=True):
        if fmt is None or value is None:
            cells.append(value)
            continue
        cell = WriteOnlyCell(ws, value=value)
        cell.number_format = _NUMBER_FORMATS[fmt]
        cells.append(cell)
    ws.append(cells)


def _total_row(ws: Worksheet, values: list, formats: list[str | None]) -> None:
    """Bold, shaded total row — values/formats align positionally with the
    sheet's header, so callers pass None for any column that isn't summed.
    """
    cells = []
    for value, fmt in zip(values, formats, strict=True):
        cell = WriteOnlyCell(ws, value=value)
        cell.font = _TOTAL_FONT
        cell.fill = _TOTAL_FILL
        if fmt is not None and value is not None:
            cell.number_format = _NUMBER_FORMATS[fmt]
        cells.append(cell)
    ws.append(cells)


async def _write_company_sheet(wb: Workbook, company: Company) -> None:
    ws = wb.create_sheet("Company")
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    meta_font = Font(bold=True)
    for label, value in [
        ("Generated At", datetime.now(UTC).isoformat()),
        ("Company", company.business_name),
        ("Generated By", "OpsGenie"),
        ("Format Version", EXPORT_FORMAT_VERSION),
    ]:
        label_cell = WriteOnlyCell(ws, value=label)
        label_cell.font = meta_font
        ws.append([label_cell, value])
    ws.append([])
    _write_header(ws, ["Field", "Value"])
    for label, value in [
        ("Business Name", company.business_name),
        ("Owner Name", company.owner_name),
        ("WhatsApp Number", company.whatsapp_number),
        ("Business Type", company.business_type),
        ("GST Number", company.gst_number),
        ("GST Rate (%)", company.gst_rate),
        ("Timezone", company.timezone),
        ("Subscription Active", company.subscription_active),
        ("Opening Balance", company.opening_balance),
        ("Briefing Hour", company.briefing_hour),
        ("Evening Brief Hour", company.evening_brief_hour),
        ("Activated/Created At", company.created_at.isoformat() if company.created_at else None),
    ]:
        ws.append([label, value])


_PRODUCTS_HEADERS = [
    "Name",
    "Unit",
    "Selling Price",
    "Purchase Price",
    "Stock Quantity",
    "Stock Value (Cost)",
    "Stock Value (Selling)",
    "Potential Profit",
]
_PRODUCTS_FORMATS = [None, None, "money", "money", "qty", "money", "money", "money"]


async def _write_products_sheet(db: AsyncSession, wb: Workbook, company_id: uuid.UUID) -> None:
    """Stock Value / Potential Profit are inventory valuation (what's sitting
    on the shelf right now), not the realized, accrual-basis sales margin
    already tracked per-day in Daily Business Summary — the two are kept
    separate rather than blended into one figure (see daily_snapshot.py).
    """
    ws = wb.create_sheet("Products")
    _write_header(ws, _PRODUCTS_HEADERS)
    products = await db.scalars(
        select(Product).where(Product.company_id == company_id).order_by(Product.name)
    )
    total_cost_value = Decimal("0")
    total_selling_value = Decimal("0")
    total_profit = Decimal("0")
    has_any = False
    for p in products:
        cost_value = (
            p.purchase_price * p.stock_quantity if p.purchase_price is not None else None
        )
        selling_value = (
            p.selling_price * p.stock_quantity if p.selling_price is not None else None
        )
        profit = (
            selling_value - cost_value
            if cost_value is not None and selling_value is not None
            else None
        )
        _row(
            ws,
            [
                p.name,
                p.unit,
                p.selling_price,
                p.purchase_price,
                p.stock_quantity,
                cost_value,
                selling_value,
                profit,
            ],
            _PRODUCTS_FORMATS,
        )
        total_cost_value += cost_value or Decimal("0")
        total_selling_value += selling_value or Decimal("0")
        total_profit += profit or Decimal("0")
        has_any = True
    if has_any:
        _total_row(
            ws,
            ["TOTAL", None, None, None, None, total_cost_value, total_selling_value, total_profit],
            _PRODUCTS_FORMATS,
        )


_PARTY_HEADERS = [
    "Name",
    "Phone",
    "Address",
    "GST Number",
    "Payment Terms (days)",
    "Credit Limit",
    "Notes",
    "Outstanding",
]
_PARTY_FORMATS = [None, None, None, None, None, "money", None, "money"]


async def _write_party_sheet(
    db: AsyncSession,
    wb: Workbook,
    company_id: uuid.UUID,
    *,
    model: type[Dealer] | type[Supplier],
    sheet_name: str,
    direction: str,
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, _PARTY_HEADERS)
    parties = (
        (await db.scalars(select(model).where(model.company_id == company_id).order_by(model.name)))
        .all()
    )
    outstanding_by_party = await calculate_outstanding_for_company(
        db, company_id=company_id, direction=direction
    )
    total_outstanding = Decimal("0")
    for party in parties:
        # Dealer has address/gst_number; Supplier doesn't — getattr keeps one
        # shared column layout for both sheets instead of two divergent ones.
        outstanding = outstanding_by_party.get(party.id, 0)
        _row(
            ws,
            [
                party.name,
                party.phone,
                getattr(party, "address", None),
                getattr(party, "gst_number", None),
                party.payment_terms_days,
                party.credit_limit,
                party.notes,
                outstanding,
            ],
            _PARTY_FORMATS,
        )
        total_outstanding += outstanding or Decimal("0")
    if parties:
        _total_row(
            ws,
            ["TOTAL", None, None, None, None, None, None, total_outstanding],
            _PARTY_FORMATS,
        )


_OPEN_INVOICES_HEADERS = [
    "Counterparty",
    "Invoice Number",
    "Due Date",
    "Days Overdue",
    "Amount Outstanding",
]
_OPEN_INVOICES_FORMATS = [None, None, "date", None, "money"]


async def _write_open_invoices_sheet(
    db: AsyncSession,
    wb: Workbook,
    company_id: uuid.UUID,
    *,
    sheet_name: str,
    direction: InvoiceDirection,
) -> None:
    ws = wb.create_sheet(sheet_name)
    _write_header(ws, _OPEN_INVOICES_HEADERS)
    today = datetime.now(UTC).date()

    invoices = (
        (
            await db.scalars(
                select(Invoice)
                .where(
                    Invoice.company_id == company_id,
                    Invoice.direction == direction,
                    Invoice.status.in_(_OPEN_STATUSES),
                )
                .order_by(Invoice.due_date)
            )
        )
        .all()
    )
    if not invoices:
        return
    invoice_ids = [inv.id for inv in invoices]
    paid_rows = await db.execute(
        select(Payment.invoice_id, func.sum(Payment.amount))
        .where(Payment.invoice_id.in_(invoice_ids))
        .group_by(Payment.invoice_id)
    )
    paid_by_invoice: dict[uuid.UUID, object] = dict(paid_rows.all())

    party_ids = {inv.dealer_id if direction == InvoiceDirection.receivable else inv.supplier_id
                 for inv in invoices}
    party_ids.discard(None)
    model = Dealer if direction == InvoiceDirection.receivable else Supplier
    names = {
        p.id: p.name
        for p in (await db.scalars(select(model).where(model.id.in_(party_ids)))).all()
    }

    total_outstanding = Decimal("0")
    for inv in invoices:
        party_id = inv.dealer_id if direction == InvoiceDirection.receivable else inv.supplier_id
        paid = paid_by_invoice.get(inv.id, 0)
        outstanding = inv.total_amount - paid
        days_overdue = max((today - inv.due_date).days, 0)
        _row(
            ws,
            [
                names.get(party_id, "Unknown"),
                inv.invoice_number,
                inv.due_date,
                days_overdue,
                outstanding,
            ],
            _OPEN_INVOICES_FORMATS,
        )
        total_outstanding += outstanding
    _total_row(
        ws,
        ["TOTAL", None, None, None, total_outstanding],
        _OPEN_INVOICES_FORMATS,
    )


_INVOICES_HEADERS = [
    "Invoice Number",
    "Direction",
    "Counterparty",
    "Invoice Date",
    "Due Date",
    "Subtotal",
    "GST Amount",
    "Total Amount",
    "Status",
    "Source",
    "Amount Paid",
    "Amount Outstanding",
]
_INVOICES_FORMATS = [
    None, None, None, "date", "date", "money", "money", "money", None, None, "money", "money",
]


async def _write_invoices_sheet(db: AsyncSession, wb: Workbook, company_id: uuid.UUID) -> None:
    ws = wb.create_sheet("Invoices")
    _write_header(ws, _INVOICES_HEADERS)
    dealers = {
        d.id: d.name
        for d in (await db.scalars(select(Dealer).where(Dealer.company_id == company_id))).all()
    }
    suppliers = {
        s.id: s.name
        for s in (await db.scalars(select(Supplier).where(Supplier.company_id == company_id))).all()
    }
    # One grouped query for every invoice's paid total — not one SUM query per
    # invoice, which would defeat the point of streaming the invoices below.
    paid_rows = await db.execute(
        select(Payment.invoice_id, func.sum(Payment.amount))
        .join(Invoice, Payment.invoice_id == Invoice.id)
        .where(Invoice.company_id == company_id)
        .group_by(Payment.invoice_id)
    )
    paid_by_invoice: dict[uuid.UUID, object] = dict(paid_rows.all())

    stmt = (
        select(Invoice)
        .where(Invoice.company_id == company_id)
        .order_by(Invoice.invoice_date)
        .execution_options(yield_per=_STREAM_CHUNK)
    )
    result = await db.stream_scalars(stmt)
    total_subtotal = Decimal("0")
    total_gst = Decimal("0")
    total_amount = Decimal("0")
    total_paid = Decimal("0")
    total_outstanding = Decimal("0")
    has_any = False
    async for invoice in result:
        paid_total = paid_by_invoice.get(invoice.id, 0)
        outstanding = invoice.total_amount - paid_total
        counterparty = (
            dealers.get(invoice.dealer_id)
            if invoice.direction == InvoiceDirection.receivable
            else suppliers.get(invoice.supplier_id)
        )
        _row(
            ws,
            [
                invoice.invoice_number,
                _s(invoice.direction),
                counterparty or "Unknown",
                invoice.invoice_date,
                invoice.due_date,
                invoice.subtotal,
                invoice.gst_amount,
                invoice.total_amount,
                _s(invoice.status),
                _s(invoice.source),
                paid_total,
                outstanding,
            ],
            _INVOICES_FORMATS,
        )
        total_subtotal += invoice.subtotal
        total_gst += invoice.gst_amount
        total_amount += invoice.total_amount
        total_paid += paid_total or Decimal("0")
        total_outstanding += outstanding
        has_any = True
    if has_any:
        _total_row(
            ws,
            [
                "TOTAL", None, None, None, None,
                total_subtotal, total_gst, total_amount, None, None,
                total_paid, total_outstanding,
            ],
            _INVOICES_FORMATS,
        )


_PAYMENTS_HEADERS = ["Amount", "Payment Date", "Method", "Source", "Invoice Number", "Direction"]
_PAYMENTS_FORMATS = ["money", "date", None, None, None, None]


async def _write_payments_sheet(db: AsyncSession, wb: Workbook, company_id: uuid.UUID) -> None:
    ws = wb.create_sheet("Payments")
    _write_header(ws, _PAYMENTS_HEADERS)
    invoice_lookup = {
        inv.id: (inv.invoice_number, inv.direction)
        for inv in (
            await db.scalars(select(Invoice).where(Invoice.company_id == company_id))
        ).all()
    }
    stmt = (
        select(Payment)
        .where(Payment.company_id == company_id)
        .order_by(Payment.payment_date)
        .execution_options(yield_per=_STREAM_CHUNK)
    )
    result = await db.stream_scalars(stmt)
    total_amount = Decimal("0")
    has_any = False
    async for payment in result:
        invoice_number, direction = invoice_lookup.get(payment.invoice_id, (None, None))
        _row(
            ws,
            [
                payment.amount,
                payment.payment_date,
                payment.method,
                _s(payment.source),
                invoice_number,
                _s(direction),
            ],
            _PAYMENTS_FORMATS,
        )
        total_amount += payment.amount
        has_any = True
    if has_any:
        _total_row(ws, [total_amount, None, None, None, None, None], _PAYMENTS_FORMATS)


_DAILY_SUMMARY_HEADERS = [
    "Date",
    "Sales",
    "Sales Margin",
    "Collections",
    "Supplier Payments",
    "Net Cash Movement",
    "Outstanding Receivables",
    "Invoices Created",
    "Payments Recorded",
    "Orders via WhatsApp",
]
_DAILY_SUMMARY_FORMATS = [
    "date", "money", "money", "money", "money", "money", "money", None, None, None,
]


async def _write_daily_summary_sheet(db: AsyncSession, wb: Workbook, company_id: uuid.UUID) -> None:
    ws = wb.create_sheet("Daily Business Summary")
    _write_header(ws, _DAILY_SUMMARY_HEADERS)
    rows = await db.scalars(
        select(DailyBusinessSnapshot)
        .where(DailyBusinessSnapshot.company_id == company_id)
        .order_by(DailyBusinessSnapshot.business_date)
    )
    total_sales = Decimal("0")
    total_margin = Decimal("0")
    total_collections = Decimal("0")
    total_supplier_payments = Decimal("0")
    total_net_cash = Decimal("0")
    has_any = False
    for r in rows:
        _row(
            ws,
            [
                r.business_date,
                r.sales_amount,
                r.sales_margin,
                r.collections_amount,
                r.supplier_payments_amount,
                r.net_cash_movement,
                r.outstanding_receivables,
                r.invoices_created,
                r.payments_recorded,
                r.orders_created,
            ],
            _DAILY_SUMMARY_FORMATS,
        )
        total_sales += r.sales_amount
        total_margin += r.sales_margin
        total_collections += r.collections_amount
        total_supplier_payments += r.supplier_payments_amount
        total_net_cash += r.net_cash_movement
        has_any = True
    if has_any:
        # Outstanding Receivables is a point-in-time balance, not a flow — it
        # is deliberately left out of the total row rather than summed across
        # days, which would produce a meaningless number (see daily_snapshot.py).
        _total_row(
            ws,
            [
                "TOTAL", total_sales, total_margin, total_collections,
                total_supplier_payments, total_net_cash, None, None, None, None,
            ],
            _DAILY_SUMMARY_FORMATS,
        )


async def _write_faqs_sheet(db: AsyncSession, wb: Workbook, company_id: uuid.UUID) -> None:
    ws = wb.create_sheet("FAQs")
    _write_header(ws, ["Question", "Answer"])
    faqs = await db.scalars(
        select(FAQ).where(FAQ.company_id == company_id).order_by(FAQ.created_at)
    )
    for faq in faqs:
        ws.append([faq.question, faq.answer])


async def build_company_workbook(db: AsyncSession, company: Company) -> bytes:
    """Rebuild the entire per-company workbook fresh from the live database.
    Pure read — never commits, never mutates. A freshly-activated company with
    no data yet naturally produces headers-only sheets.
    """
    wb = Workbook(write_only=True)
    await _write_company_sheet(wb, company)
    await _write_products_sheet(db, wb, company.id)
    await _write_party_sheet(
        db, wb, company.id, model=Dealer, sheet_name="Dealers", direction="receivable"
    )
    await _write_party_sheet(
        db, wb, company.id, model=Supplier, sheet_name="Suppliers", direction="payable"
    )
    await _write_open_invoices_sheet(
        db, wb, company.id, sheet_name="Receivables", direction=InvoiceDirection.receivable
    )
    await _write_open_invoices_sheet(
        db, wb, company.id, sheet_name="Payables", direction=InvoiceDirection.payable
    )
    await _write_invoices_sheet(db, wb, company.id)
    await _write_payments_sheet(db, wb, company.id)
    await _write_daily_summary_sheet(db, wb, company.id)
    await _write_faqs_sheet(db, wb, company.id)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_export_link(company: Company, *, base_url: str) -> str:
    """Stateless short-lived signed link — nothing is written to the database
    to "issue" it, so there's nothing to store or revoke. Same HMAC signing
    convention already used to verify Meta's webhook payloads.
    """
    settings = get_settings()
    if not settings.export_link_secret:
        raise ValueError("EXPORT_LINK_SECRET is not configured.")
    expires_at = int(time.time()) + settings.export_link_ttl_minutes * 60
    message = f"{company.id}:{expires_at}"
    signature = hmac.new(
        settings.export_link_secret.encode(), message.encode(), hashlib.sha256
    ).hexdigest()
    return f"{base_url.rstrip('/')}/export/{company.id}/{expires_at}/{signature}"


def verify_export_link(company_id: uuid.UUID, expires_at: int, signature: str) -> bool:
    settings = get_settings()
    if not settings.export_link_secret:
        return False
    if time.time() > expires_at:
        return False
    message = f"{company_id}:{expires_at}"
    expected = hmac.new(
        settings.export_link_secret.encode(), message.encode(), hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(expected, signature)
